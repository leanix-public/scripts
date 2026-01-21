#!/usr/bin/env python3
"""
Standalone script to query LeanIX Synchronization Items API with status filtering.

API Documentation:
https://eu.leanix.net/openapi-explorer/?urls.primaryName=Synclog

Usage:
    python query_sync_items_with_status.py --api-token TOKEN --workspace-id <YOUR_WORKSPACE_ID> --region us --status OK
    python query_sync_items_with_status.py --api-token TOKEN --workspace-id <YOUR_WORKSPACE_ID> --region orgname --start-date 2026-01-01
"""

import asyncio
import argparse
import json
import logging
import os
import sys
import time
from typing import Optional, Dict, Any, List, AsyncIterator
from datetime import datetime, timezone, timedelta

try:
    from dateutil import parser as date_parser
except ImportError:
    date_parser = None

try:
    import httpx
except ImportError:
    print("Error: httpx library not found. Install it with: pip install httpx")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def setup_logging(debug: bool = False) -> None:
    """Configure logging to avoid sensitive data exposure."""
    level = logging.DEBUG if debug else logging.INFO
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[logging.StreamHandler()]
    )


def parse_iso_datetime(date_str: str) -> datetime:
    """Parse ISO datetime string to datetime object."""
    return datetime.fromisoformat(date_str.replace('Z', '+00:00'))


def format_iso_datetime(dt: datetime) -> str:
    """Format datetime object to ISO string."""
    return dt.isoformat().replace('+00:00', 'Z')


def mask_sensitive_data(data: str, max_length: int = 100) -> str:
    """Mask potentially sensitive data for logging."""
    if not data:
        return "[empty]"
    if len(data) <= 20:
        return "***"
    return f"{data[:8]}...{data[-4:]}"


class LeanIXSyncItemsClient:
    """Client for querying LeanIX Synchronization Items with status filtering."""
    
    VALID_STATUSES = ["OK", "WARNING", "ERROR", "INFO"]
    
    def __init__(self, api_token: str, region: str, workspace_id: Optional[str] = None):
        self.api_token = api_token
        self.region = region
        self.workspace_id = workspace_id
        self.auth_url = f"https://{region}.leanix.net/services/mtm/v1/oauth2/token"
        self.base_url = f"https://{region}.leanix.net/services/synclog/v1"
        self.bearer_token: Optional[str] = None
        self.token_expires_at: Optional[float] = None
        self._client: Optional[httpx.AsyncClient] = None
        
        # Log initialization with masked token
        masked_token = mask_sensitive_data(api_token)
        logging.info(f"Initializing client with token: {masked_token}, region: {region}")
    
    async def __aenter__(self):
        """Async context manager entry."""
        self._client = httpx.AsyncClient(
            timeout=httpx.Timeout(60.0),
            limits=httpx.Limits(max_keepalive_connections=5, max_connections=10),
        )
        return self
    
    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Async context manager exit."""
        if self._client:
            await self._client.aclose()
            
    async def get_bearer_token(self) -> str:
        """Exchange API token for Bearer token with retry logic."""
        if self.bearer_token and self.token_expires_at:
            if time.time() < self.token_expires_at - 60:
                return self.bearer_token
        
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {"grant_type": "client_credentials"}
        auth = ("apitoken", self.api_token)
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                logging.info(f"🔐 Authenticating with LeanIX ({self.region})... (attempt {attempt + 1})")
                response = await self._client.post(self.auth_url, headers=headers, data=data, auth=auth)
                response.raise_for_status()
                
                token_data = response.json()
                self.bearer_token = token_data.get("access_token")
                expires_in = token_data.get("expires_in", 3600)
                
                if not self.bearer_token:
                    raise ValueError("No access_token in response")
                
                self.token_expires_at = time.time() + expires_in
                
                logging.info(f"✓ Successfully obtained Bearer token (expires in {expires_in}s)")
                return self.bearer_token
                
            except httpx.HTTPStatusError as e:
                if e.response.status_code == 429:  # Rate limited
                    retry_after = int(e.response.headers.get("Retry-After", 2 ** attempt))
                    logging.warning(f"Rate limited, waiting {retry_after}s before retry {attempt+1}/{max_retries}")
                    await asyncio.sleep(retry_after)
                    continue
                elif e.response.status_code in [500, 502, 503, 504] and attempt < max_retries - 1:
                    wait_time = 2 ** attempt
                    logging.warning(f"Server error {e.response.status_code}, retrying in {wait_time}s")
                    await asyncio.sleep(wait_time)
                    continue
                else:
                    logging.error(f"✗ Authentication failed: {e.response.status_code}")
                    # Only show response details in debug mode
                    if logging.getLogger().level == logging.DEBUG:
                        logging.debug(f"Response: {e.response.text}")
                    else:
                        logging.error("Response: [Hidden - set --debug to see details]")
                    raise
            except httpx.RequestError as e:
                if attempt < max_retries - 1:
                    wait_time = 2 ** attempt
                    logging.warning(f"Connection error, retrying in {wait_time}s: {type(e).__name__}")
                    await asyncio.sleep(wait_time)
                    continue
                else:
                    logging.error(f"✗ Authentication connection error: {e}")
                    raise
            except Exception as e:
                logging.error(f"✗ Authentication error: {type(e).__name__}: {e}")
                raise
        
        raise RuntimeError("Failed to authenticate after all retries")
    
    async def query_sync_items(self, status: Optional[str] = None, limit: Optional[int] = None, 
                               cursor: Optional[str] = None, page: int = 1) -> Dict[str, Any]:
        """Query the /syncItems endpoint with optional status filter."""
        bearer_token = await self.get_bearer_token()
        
        headers = {
            "Authorization": f"Bearer {bearer_token}",
            "Content-Type": "application/json",
            "Accept": "application/json"
        }
        
        # Build query parameters to match Swagger UI format
        params = {}
        
        # Add workspace ID as query parameter (not header)
        if self.workspace_id:
            params["workspaceId"] = self.workspace_id
        
        if status:
            if status.upper() not in self.VALID_STATUSES:
                logging.warning(f"⚠ Warning: '{status}' may not be a valid status. Valid: {', '.join(self.VALID_STATUSES)}")
            params["status"] = status.upper()
        
        # Use 'size' instead of 'limit' to match Swagger UI
        if limit:
            params["size"] = limit
        else:
            params["size"] = 30  # Default size from Swagger UI
        
        # Add page parameter
        params["page"] = page
        
        # Add sorting (newest first, matching Swagger UI)
        params["sort"] = "createdAt-desc"
        
        # Add cursor for pagination if provided
        if cursor:
            params["cursor"] = cursor
        
        url = f"{self.base_url}/syncItems"
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                logging.debug(f"\n🔍 Querying {url}")
                if params:
                    logging.debug(f"   Parameters: {json.dumps(params, indent=2)}")
                
                response = await self._client.get(url, headers=headers, params=params)
                response.raise_for_status()
                
                data = response.json()
                logging.info(f"✓ Successfully retrieved sync items (page size: {len(data.get('data', []))})")
                
                return data
                
            except httpx.HTTPStatusError as e:
                if e.response.status_code == 429:  # Rate limited
                    retry_after = int(e.response.headers.get("Retry-After", 2 ** attempt))
                    logging.warning(f"Rate limited, waiting {retry_after}s before retry {attempt+1}/{max_retries}")
                    await asyncio.sleep(retry_after)
                    continue
                elif e.response.status_code in [500, 502, 503, 504] and attempt < max_retries - 1:
                    wait_time = 2 ** attempt
                    logging.warning(f"Server error {e.response.status_code}, retrying in {wait_time}s")
                    await asyncio.sleep(wait_time)
                    continue
                else:
                    logging.error(f"✗ Request failed: {e.response.status_code}")
                    # Only show response details in debug mode
                    if logging.getLogger().level == logging.DEBUG:
                        logging.debug(f"Response: {e.response.text}")
                    else:
                        logging.error("Response: [Hidden - set --debug to see details]")
                    raise
            except httpx.RequestError as e:
                if attempt < max_retries - 1:
                    wait_time = 2 ** attempt
                    logging.warning(f"Connection error, retrying in {wait_time}s: {type(e).__name__}")
                    await asyncio.sleep(wait_time)
                    continue
                else:
                    logging.error(f"✗ Request connection error: {e}")
                    raise
            except Exception as e:
                logging.error(f"✗ Request error: {type(e).__name__}: {e}")
                raise
        
        raise RuntimeError("Failed to complete request after all retries")
    
    async def get_all_sync_items(self, status: Optional[str] = None, batch_limit: int = 30,
                                 max_items: Optional[int] = None, start_date: Optional[datetime] = None,
                                 end_date: Optional[datetime] = None) -> List[Dict[str, Any]]:
        """Get all sync items with automatic pagination and optional client-side date filtering."""
        all_items = []
        page = 1
        
        while True:
            logging.info(f"\n📄 Fetching page {page}...")
            
            response = await self.query_sync_items(status=status, limit=batch_limit, page=page)
            
            items = response.get("data", [])
            if not items:
                logging.info("   No more items found")
                break
            
            # Filter items by date range if specified
            filtered_items = []
            items_before_range = 0
            items_after_range = 0
            
            for item in items:
                created_at_str = item.get("createdAt", "")
                if not created_at_str:
                    continue
                
                try:
                    created_at = parse_iso_datetime(created_at_str)
                    
                    # Proper datetime comparison
                    if start_date and created_at < start_date:
                        items_before_range += 1
                        continue
                    if end_date and created_at > end_date:
                        items_after_range += 1
                        continue
                    
                    filtered_items.append(item)
                except ValueError:
                    logging.warning(f"Invalid date format in item: {created_at_str}")
                    continue
            
            # If all items are before start_date, we can stop (assumes items are sorted newest first)
            if start_date and items_before_range == len(items):
                logging.info(f"   All {len(items)} items are before start date - stopping pagination")
                break
            
            all_items.extend(filtered_items)
            logging.info(f"   Retrieved {len(items)} items, {len(filtered_items)} match date range (total: {len(all_items)})")
            
            if max_items and len(all_items) >= max_items:
                all_items = all_items[:max_items]
                logging.info(f"   Reached maximum of {max_items} items")
                break
            
            # Check if we got fewer items than requested, indicating last page
            if len(items) < batch_limit:
                logging.info("   Reached last page (fewer items than batch size)")
                break
            
            page += 1
            
            if page > 100:
                logging.warning("   Reached safety limit of 100 pages")
                break
        
        return all_items

    async def stream_sync_items(self, status: Optional[str] = None, batch_limit: int = 30,
                                start_date: Optional[datetime] = None,
                                end_date: Optional[datetime] = None) -> AsyncIterator[Dict[str, Any]]:
        """Stream sync items to avoid loading all in memory."""
        page = 1
        
        while True:
            logging.info(f"\n📄 Streaming page {page}...")
            
            response = await self.query_sync_items(status=status, limit=batch_limit, page=page)
            
            items = response.get("data", [])
            if not items:
                break
            
            items_yielded = 0
            for item in items:
                created_at_str = item.get("createdAt", "")
                if not created_at_str:
                    continue
                
                try:
                    created_at = parse_iso_datetime(created_at_str)
                    
                    if start_date and created_at < start_date:
                        continue
                    if end_date and created_at > end_date:
                        continue
                    
                    yield item
                    items_yielded += 1
                except ValueError:
                    logging.warning(f"Invalid date format in item: {created_at_str}")
                    continue
            
            logging.info(f"   Processed {len(items)} items, yielded {items_yielded}")
            
            # Check if we got fewer items than requested, indicating last page
            if len(items) < batch_limit:
                logging.info("   Reached last page (fewer items than batch size)")
                break
            
            page += 1
            if page > 100:
                logging.warning("   Reached safety limit of 100 pages")
                break


def print_summary(items: List[Dict[str, Any]], status_filter: Optional[str] = None) -> None:
    """Print a summary of sync items."""
    print("\n" + "="*80)
    print(f"SYNC ITEMS SUMMARY")
    if status_filter:
        print(f"Filter: status = {status_filter.upper()}")
    print("="*80)
    print(f"Total items: {len(items)}")
    
    if not items:
        print("No items found.")
        return
    
    types = {}
    statuses = {}
    sources = {}
    
    for item in items:
        item_type = item.get("type", "unknown")
        status = item.get("status", "unknown")
        source = item.get("source", "unknown")
        
        types[item_type] = types.get(item_type, 0) + 1
        statuses[status] = statuses.get(status, 0) + 1
        sources[source] = sources.get(source, 0) + 1
    
    if types:
        print("\nItems by type:")
        for item_type, count in sorted(types.items()):
            print(f"  {item_type}: {count}")
    
    if statuses:
        print("\nItems by status:")
        for status, count in sorted(statuses.items()):
            print(f"  {status}: {count}")
    
    if sources:
        print("\nItems by source:")
        for source, count in sorted(sources.items()):
            print(f"  {source}: {count}")
    
    print("\n" + "="*80)


def save_to_file(items: List[Dict[str, Any]], filename: str, status_filter: Optional[str] = None) -> None:
    """Save sync items to a JSON file."""
    output = {
        "timestamp": datetime.now().isoformat(),
        "status_filter": status_filter,
        "total_items": len(items),
        "items": items
    }
    
    try:
        with open(filename, 'w') as f:
            json.dump(output, f, indent=2)
        
        print(f"\n💾 Saved {len(items)} items to {filename}")
    except Exception as e:
        logging.error(f"Failed to save to file {filename}: {e}")
        raise


def save_to_excel(items: List[Dict[str, Any]], filename: str, status_filter: Optional[str] = None, 
                  start_date: Optional[datetime] = None, end_date: Optional[datetime] = None, 
                  date_range_defaulted: bool = False) -> None:
    """Save sync items to an Excel file with formatting."""
    if not EXCEL_AVAILABLE:
        logging.error("Excel export not available. Install openpyxl with: pip install openpyxl")
        return
    
    try:
        wb = Workbook()
        
        # Create main data worksheet
        ws = wb.active
        ws.title = "Sync Items"
        
        if not items:
            ws['A1'] = "No sync items found"
            wb.save(filename)
            print(f"\n📊 Saved empty Excel file to {filename}")
            return
        
        # Get all unique keys from all items to create comprehensive headers
        all_keys = set()
        for item in items:
            all_keys.update(item.keys())
        
        # Define preferred column order
        preferred_order = [
            'id', 'status', 'type', 'source', 'createdAt', 'updatedAt',
            'startedAt', 'finishedAt', 'message', 'errorMessage', 'workspaceId'
        ]
        
        # Order headers: preferred first, then remaining alphabetically
        headers = []
        for key in preferred_order:
            if key in all_keys:
                headers.append(key)
                all_keys.remove(key)
        headers.extend(sorted(all_keys))
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows
        for row_idx, item in enumerate(items, 2):
            for col_idx, header in enumerate(headers, 1):
                value = item.get(header, "")
                
                # Format datetime values
                if header in ['createdAt', 'updatedAt', 'startedAt', 'finishedAt'] and value:
                    try:
                        dt = parse_iso_datetime(value)
                        value = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass  # Keep original value if parsing fails
                
                # Handle nested objects/arrays
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Color code by status
                if header == 'status':
                    if value == 'SUCCESS':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                    elif value == 'FAILED':
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                    elif value == 'RUNNING':
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, max 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary worksheet
        summary_ws = wb.create_sheet("Summary")
        
        # Summary data
        summary_data = [
            ["Report Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Status Filter", status_filter if status_filter else "All"],
            ["Total Items", len(items)],
        ]
        
        # Add date range information
        if start_date:
            start_label = "Start Date"
            if date_range_defaulted:
                start_label += " (Default)"
            summary_data.append([start_label, start_date.strftime('%Y-%m-%d %H:%M:%S %Z')])
        
        if end_date:
            end_label = "End Date"
            if date_range_defaulted:
                end_label += " (Default)"
            summary_data.append([end_label, end_date.strftime('%Y-%m-%d %H:%M:%S %Z')])
        
        if date_range_defaulted:
            summary_data.append(["Date Range Note", "Last 30 days (auto-applied)"])
        
        summary_data.extend([
            ["", ""],  # Empty row
            ["Status Breakdown", "Count"]
        ])
        
        # Count by status
        status_counts = {}
        type_counts = {}
        source_counts = {}
        
        for item in items:
            status = item.get('status', 'unknown')
            item_type = item.get('type', 'unknown')
            source = item.get('source', 'unknown')
            
            status_counts[status] = status_counts.get(status, 0) + 1
            type_counts[item_type] = type_counts.get(item_type, 0) + 1
            source_counts[source] = source_counts.get(source, 0) + 1
        
        # Add status counts
        for status, count in sorted(status_counts.items()):
            summary_data.append([status, count])
        
        summary_data.extend([
            ["", ""],  # Empty row
            ["Type Breakdown", "Count"]
        ])
        
        # Add type counts
        for item_type, count in sorted(type_counts.items()):
            summary_data.append([item_type, count])
        
        summary_data.extend([
            ["", ""],  # Empty row
            ["Source Breakdown", "Count"]
        ])
        
        # Add source counts
        for source, count in sorted(source_counts.items()):
            summary_data.append([source, count])
        
        # Write summary data
        for row_idx, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_idx, column=1, value=label)
            summary_ws.cell(row=row_idx, column=2, value=value)
            
            # Format headers
            if label in ["Report Generated", "Status Breakdown", "Type Breakdown", "Source Breakdown"]:
                summary_ws.cell(row=row_idx, column=1).font = Font(bold=True)
                summary_ws.cell(row=row_idx, column=2).font = Font(bold=True)
        
        # Auto-adjust summary column widths
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 15
        
        wb.save(filename)
        print(f"\n📊 Saved {len(items)} items to Excel file: {filename}")
        print(f"   Worksheets: 'Sync Items' (data) and 'Summary' (statistics)")
        
    except Exception as e:
        logging.error(f"Failed to save Excel file {filename}: {e}")
        raise


async def main():
    """Main function."""
    parser = argparse.ArgumentParser(
        description="Query LeanIX Synchronization Items API with status filtering",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Get ERROR items (automatically saves to Excel)
  python query_sync_items_with_status.py --api-token <YOUR_API_TOKEN> --workspace-id <YOUR_WORKSPACE_ID> --region <YOUR_REGION> --status ERROR
  
  # Get items with custom Excel filename
  python query_sync_items_with_status.py --api-token <YOUR_API_TOKEN> --workspace-id <YOUR_WORKSPACE_ID> --region <YOUR_REGION> --status OK --excel ok_items.xlsx
  
  # Get items with date filtering (auto Excel export)
  python query_sync_items_with_status.py --api-token <YOUR_API_TOKEN> --workspace-id <YOUR_WORKSPACE_ID> --region <YOUR_REGION> --start-date 2024-01-01 --end-date 2024-01-31
  
  # Export to JSON instead of Excel
  python query_sync_items_with_status.py --api-token <YOUR_API_TOKEN> --workspace-id <YOUR_WORKSPACE_ID> --region <YOUR_REGION> --status WARNING --output warning_items.json
  
  # Stream large datasets to Excel
  python query_sync_items_with_status.py --api-token <YOUR_API_TOKEN> --workspace-id <YOUR_WORKSPACE_ID> --region <YOUR_REGION> --stream --excel large_dataset.xlsx

Note: All three parameters are mandatory (--api-token, --workspace-id, --region)
      Or set environment variables: LEANIX_API_TOKEN, LEANIX_WORKSPACE_ID, LEANIX_REGION

Default behavior: Automatically exports to Excel file with timestamp for last 30 days
Valid status values: OK, WARNING, ERROR, INFO
        """
    )
    parser.add_argument("--api-token", help="LeanIX API token (or set LEANIX_API_TOKEN env var)")
    parser.add_argument("--workspace-id", help="LeanIX workspace ID (or set LEANIX_WORKSPACE_ID env var)")
    parser.add_argument("--region",  help="LeanIX region/sub-domain from your tenant URL (e.g., eu, demo-eu-2, us-1, orgname)")
    parser.add_argument("--status", help="Filter by status (OK, WARNING, ERROR, INFO)")
    parser.add_argument("--start-date", help="Filter by start date (YYYY-MM-DD or ISO format)")
    parser.add_argument("--end-date", help="Filter by end date (YYYY-MM-DD or ISO format)")
    parser.add_argument("--limit", type=int, help="Limit total number of items to fetch")
    parser.add_argument("--batch-size", type=int, default=100, help="Items per request (default: 100)")
    parser.add_argument("--output", help="Output JSON file path")
    parser.add_argument("--excel", help="Output Excel file path (.xlsx)")
    parser.add_argument("--stream", action="store_true", help="Stream items to handle large datasets efficiently")
    parser.add_argument("--verbose", action="store_true", help="Show detailed item information")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging (may expose sensitive data)")
    
    # Check if no arguments provided at all (just script name)
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(0)
    
    args = parser.parse_args()
    
    # Setup logging
    setup_logging(debug=args.debug)
    
    # Validate required parameters
    api_token = args.api_token or os.getenv("LEANIX_API_TOKEN")
    if not api_token:
        logging.error("✗ Error: API token is required")
        logging.error("   Use --api-token TOKEN or set LEANIX_API_TOKEN environment variable")
        parser.print_help()
        sys.exit(1)
    
    workspace_id = args.workspace_id or os.getenv("LEANIX_WORKSPACE_ID")
    if not workspace_id:
        logging.error("✗ Error: Workspace ID is required")
        logging.error("   Use --workspace-id <YOUR_WORKSPACE_ID> or set LEANIX_WORKSPACE_ID environment variable")
        parser.print_help()
        sys.exit(1)
    
    region = args.region or os.getenv("LEANIX_REGION")
    if not region:
        logging.error("✗ Error: Region is required")
        logging.error("   Use --region REGION (e.g., eu, enbridgeca, demo-eu-2, us-1)")
        logging.error("   Or set LEANIX_REGION environment variable")
        parser.print_help()
        sys.exit(1)
    
    # Parse and validate dates with default range of 1 month from current date
    start_date = None
    end_date = None
    date_range_defaulted = False
    
    # If no date range specified, default to last month
    if not args.start_date and not args.end_date:
        now = datetime.now()
        ct_timezone = timezone(timedelta(hours=-6))
        end_date = now.replace(tzinfo=ct_timezone)
        start_date = end_date - timedelta(days=30)  # 30 days ago
        date_range_defaulted = True
        logging.info(f"ℹ️  No date range specified - defaulting to last 30 days")
    
    if args.start_date:
        try:
            # Try to parse the date
            if date_parser:
                parsed_date = date_parser.parse(args.start_date)
            else:
                # Fallback to basic parsing
                parsed_date = datetime.fromisoformat(args.start_date.replace('Z', '+00:00'))
            
            # Ensure timezone-aware datetime (default to Central Time if naive)
            if parsed_date.tzinfo is None:
                # Central Time: UTC-6 (CST) or UTC-5 (CDT) - using CST as default
                ct_timezone = timezone(timedelta(hours=-6))
                parsed_date = parsed_date.replace(tzinfo=ct_timezone)
            start_date = parsed_date
        except Exception as e:
            logging.error(f"✗ Error: Invalid start date format: {args.start_date}")
            logging.error(f"   Use format: YYYY-MM-DD or ISO datetime")
            sys.exit(1)
    
    if args.end_date:
        try:
            if date_parser:
                parsed_date = date_parser.parse(args.end_date)
            else:
                parsed_date = datetime.fromisoformat(args.end_date.replace('Z', '+00:00'))
            
            # Ensure timezone-aware datetime (default to Central Time if naive)
            if parsed_date.tzinfo is None:
                # Central Time: UTC-6 (CST) or UTC-5 (CDT) - using CST as default
                ct_timezone = timezone(timedelta(hours=-6))
                parsed_date = parsed_date.replace(tzinfo=ct_timezone)
            end_date = parsed_date
        except Exception as e:
            logging.error(f"✗ Error: Invalid end date format: {args.end_date}")
            logging.error(f"   Use format: YYYY-MM-DD or ISO datetime")
            sys.exit(1)
    
    print("="*80)
    print("LeanIX Synchronization Items Query")
    print("="*80)
    print(f"Region: {region}")
    print(f"Workspace ID: {workspace_id}")
    print(f"Batch size: {args.batch_size}")
    if args.status:
        print(f"Status filter: {args.status.upper()}")
    if start_date:
        date_range_text = f"Start date: {format_iso_datetime(start_date)}"
        if date_range_defaulted:
            date_range_text += " (default: last 30 days)"
        print(date_range_text)
    if end_date:
        date_range_text = f"End date: {format_iso_datetime(end_date)}"
        if date_range_defaulted:
            date_range_text += " (default: current time)"
        print(date_range_text)
    if args.limit:
        print(f"Max items: {args.limit}")
    if args.stream:
        print("Mode: Streaming (memory efficient)")
    print("="*80)
    
    try:
        async with LeanIXSyncItemsClient(api_token, region, workspace_id) as client:
            if args.stream:
                # Stream mode for large datasets
                items = []
                item_count = 0
                async for item in client.stream_sync_items(
                    status=args.status,
                    batch_limit=args.batch_size,
                    start_date=start_date,
                    end_date=end_date
                ):
                    items.append(item)
                    item_count += 1
                    
                    if args.limit and item_count >= args.limit:
                        break
                        
                    # For streaming mode with output, collect items but log progress
                    if item_count % 1000 == 0:
                        logging.info(f"Streamed {item_count} items...")
            else:
                # Standard mode
                items = await client.get_all_sync_items(
                    status=args.status,
                    batch_limit=args.batch_size,
                    max_items=args.limit,
                    start_date=start_date,
                    end_date=end_date
                )
            
            print_summary(items, args.status)
            
            # Handle output options - Excel is default
            saved_files = []
            
            if args.output:
                save_to_file(items, args.output, args.status)
                saved_files.append(args.output)
            
            if args.excel:
                if not EXCEL_AVAILABLE:
                    logging.error("Excel export requires openpyxl. Install with: pip install openpyxl")
                else:
                    save_to_excel(items, args.excel, args.status, start_date, end_date, date_range_defaulted)
                    saved_files.append(args.excel)
            
            # Default Excel export if no output files specified
            if not args.output and not args.excel and items:
                # Generate default filename
                status_suffix = f"_{args.status.lower()}" if args.status else "_all"
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                default_excel_filename = f"leanix_sync_items{status_suffix}_{timestamp}.xlsx"
                
                if EXCEL_AVAILABLE:
                    save_to_excel(items, default_excel_filename, args.status, start_date, end_date, date_range_defaulted)
                    saved_files.append(default_excel_filename)
                else:
                    # Fallback to JSON if Excel not available
                    default_json_filename = f"leanix_sync_items{status_suffix}_{timestamp}.json"
                    save_to_file(items, default_json_filename, args.status)
                    saved_files.append(default_json_filename)
                    print("ℹ️  Note: Install openpyxl for Excel export (pip install openpyxl)")
            
            # Show verbose output if requested and no custom files were specified
            if not saved_files and items and args.verbose:
                print("\n" + "="*80)
                print("SAMPLE ITEMS (first 3)")
                print("="*80)
                for i, item in enumerate(items[:3], 1):
                    print(f"\nItem {i}:")
                    print(json.dumps(item, indent=2))
            
            print("\n✓ Query completed successfully")
        
    except KeyboardInterrupt:
        print("\n\n⚠ Interrupted by user")
        sys.exit(1)
    except httpx.HTTPStatusError as e:
        logging.error(f"\n✗ HTTP Error {e.response.status_code}: {e.response.reason_phrase}")
        sys.exit(1)
    except httpx.RequestError as e:
        logging.error(f"\n✗ Connection Error: {type(e).__name__}: {e}")
        sys.exit(1)
    except ValueError as e:
        logging.error(f"\n✗ Configuration Error: {e}")
        sys.exit(1)
    except Exception as e:
        logging.error(f"\n✗ Unexpected Error: {type(e).__name__}: {e}")
        if args.debug:
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
